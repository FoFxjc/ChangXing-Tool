#!/usr/bin/env python
# -*- coding: utf-8 -*-

import copy
import re

from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Side, Border
from openpyxl.utils import get_column_letter
from openpyxl.utils.exceptions import InvalidFileException

from utils import basic


def load(path, read_only=False, data_only=False, must_exist=False):
    """ 加载Excel文件
    :param path: <str> 要加载的Excel文件的地址路径
    :param read_only: <bool> 是否开启只读模式(默认为False)
    :param data_only: <bool> 是否只读取数据,即只提取公式结果而不提取公式(默认为False)
    :param must_exist: <bool> 在Excel文件不存在/不是Excel文件时的处理方案:True=退出程序,False=返回None(默认为True)
    :return: <openpyxl.workbook.workbook.Workbook> Openpyxl的Excel文件对象
    """
    try:
        return load_workbook(path, read_only=read_only, data_only=data_only)
    except FileNotFoundError:
        if must_exist:
            basic.sys_exit("[Error] 未找到Excel文件(" + path + ")")
        else:
            print("[Warning] 未找到Excel文件(" + path + ")")
    except InvalidFileException:
        if must_exist:
            basic.sys_exit("[Error] 目标文件不是Excel文件(" + path + ")")
        else:
            print("[Warning] 目标文件不是Excel文件(" + path + ")")


def load_sheet(path, sheet_name=None, read_only=False, data_only=False, must_exist=True):
    """ 加载指定Excel文件中的指定Sheet
    :param path: <str> 要加载的Excel文件的地址路径
    :param sheet_name: <str/None> 要加载Sheet的名称,若为None则返回当前active的Sheet
    :param read_only: <bool> 是否开启只读模式(默认为False)
    :param data_only: <bool> 是否只读取数据,即只提取公式结果而不提取公式(默认为False)
    :param must_exist: <bool> 在Excel文件及Sheet不存在/不是Excel文件时的处理方案:True=退出程序,False=返回None(默认为True)
    :return: <openpyxl.worksheet.worksheet.Worksheet> Openpyxl的Sheet对象
    """
    excel = load(path, read_only=read_only, data_only=data_only, must_exist=must_exist)
    if excel is None:
        return None
    if sheet_name is None:
        result = copy.deepcopy(excel.active)
        excel.close()
        return result
    if sheet_name in excel.sheetnames:
        result = copy.deepcopy(excel[sheet_name])
        excel.close()
        return result
    else:
        excel.close()
        if must_exist:
            basic.sys_exit("[Error] 未在Excel文件中找到对应Sheet(" + path + "," + sheet_name + ")")
        else:
            print("[Warning] 未在Excel文件中找到对应Sheet(" + path + ":" + sheet_name + ")")


def check_sheet(excel, name_list):
    """ 检查Excel文件中是否包含指定的Sheet
    :param excel: <openpyxl.workbook.workbook.Workbook> 要写入文件的Openpyxl的Excel文件对象
    :param name_list: <list> 需要检查是否存在的Sheet名称列表
    :return: <bool> 是否包含所有需要存在的Sheet
    """
    for name in name_list:
        if name not in excel.sheetnames:
            basic.sys_exit("[Error] 在Excel文件中未找到需要的Sheet: " + name + " !")


def compare_sheet(sheet_1, sheet_2, compare_row=True, compare_column=True, title_rn=None):
    """ 比较两个Sheet是否相似
    :param sheet_1: <openpyxl.workbook.workbook.Workbook> 需要比较的第1个Sheet
    :param sheet_2: <openpyxl.workbook.workbook.Workbook> 需要比较的第2个Sheet
    :param compare_row: <bool> 是否比较Sheet的总行数(默认为True)
    :param compare_column: <bool> 是否比较Sheet的总列数(默认为True)
    :param title_rn: <int/None> Sheet中标题行的行数(若为None则不比较标题行)(默认为None)
    :return: <bool> 两个Sheet是否相似(相似=True,不相似=False)
    """
    if compare_row and sheet_1.max_row != sheet_2.max_row:
        print("[Info] 表格 " + sheet_1.title + " 与表格 " + sheet_2.title + " 总行数不同...")
        return False
    if compare_column and sheet_1.max_column != sheet_2.max_column:
        print("[Info] 表格 " + sheet_1.title + " 与表格 " + sheet_2.title + " 总列数不同...")
        return False
    if title_rn is not None and isinstance(title_rn, int):
        for j in range(1, min(sheet_1.max_column, sheet_2.max_column)):
            if cell_value(sheet_1, row=title_rn, column=j) != cell_value(sheet_2, row=title_rn, column=j):
                return False


def create_with_sheet(sheet_list):
    """ 创建一个空Excel工作簿并包含指定Sheet
    :param sheet_list: <list> 创建的工作簿包含的Sheet名称列表
    :return: <openpyxl.workbook.workbook.Workbook> 创建完成的空工作簿
    """
    result = Workbook()
    arrange_sheet(result, sheet_list)
    return result


def arrange_sheet(excel, sheet_list):
    """ 整理Excel文件的Sheet,删除多余的Sheet,新建缺少的Sheet(不能实现排序)
    :param excel: <openpyxl.workbook.workbook.Workbook> 需要整理Sheet的Excel文件对象
    :param sheet_list: <list> 目标Sheet列表
    :return <None>
    """
    if sheet_list is not None and sheet_list != []:
        for name in sheet_list:
            if name not in excel.sheetnames:
                excel.create_sheet(name)
        delete_list = []
        for name in excel.sheetnames:
            if name not in sheet_list:
                delete_list.append(name)
        for name in delete_list:
            excel.remove_sheet(excel[name])


def save(excel, path, must_write=True):
    """ 将Openpyxl的Excel文件对象写入到文件中
    :param excel: <openpyxl.workbook.workbook.Workbook> 要写入文件的Openpyxl的Excel文件对象
    :param path: <str> 要写入的文件地址路径
    :param must_write: <bool> 在写入失败时(数据文件被占用)的处理方案:True=退出程序,False=返回False(默认为True)
    :return: <bool> 写入是否成功
    """
    try:
        excel.save(path)
        return True
    except PermissionError:
        if must_write:
            basic.sys_exit("结果数据文件被占用，写出到文件失败!")
        else:
            return False


def find_column(sheet, column_title, row_n=1, start_col_n=1, end_col_n=65536, must_exist=True):
    """ 在Sheet中查找指定标题的列,并返回列坐标
    :param sheet: <openpyxl.worksheet.worksheet.Worksheet> 需要被查询列的Sheet对象
    :param column_title: <str> 目标列的标题
    :param row_n: <int> 列的标题存在的行坐标
    :param start_col_n: <int> 查询列标题的范围(从第多少列开始查询)
    :param end_col_n: <int> 查询列标题的范围(查询到多少列为止)
    :param must_exist: <bool> 在目标列不存在时的处理方案:True=退出程序,False=返回None(默认为True)
    :return: <int> Sheet中目标列的列坐标
    """
    for j in range(max(1, start_col_n), min(sheet.max_column, end_col_n) + 1):
        if sheet.cell(row=row_n, column=j).value == column_title:
            return j
    if must_exist:
        basic.sys_exit("[Error] 目标列在 Sheet:" + sheet.title + " 中不存在(列名:" + column_title + ")")


def find_some_column(sheet, column_title_list, row_n=1, start_col_n=1, end_col_n=65536, must_exist=True):
    """ 在Sheet中查找一组指定标题的列,并返回一组列坐标(Dict格式)
    :param sheet: <openpyxl.worksheet.worksheet.Worksheet> 需要被查询列的Sheet对象
    :param column_title_list: <str> 目标列的标题
    :param row_n: <int> 列的标题存在的行坐标
    :param start_col_n: <int> 查询列标题的范围(从第多少列开始查询)
    :param end_col_n: <int> 查询列标题的范围(查询到多少列为止)
    :param must_exist: <bool> 在目标列不存在时的处理方案:True=退出程序,False=返回None(默认为True)
    :return: <dict> Sheet中该组目标列的列坐标字典,例如: {'平台':1, '目前名称':2}
    """
    col_n_list = {}
    for column_title in column_title_list:
        col_n_list[column_title] = find_column(sheet, column_title, row_n, start_col_n, end_col_n, must_exist)
    return col_n_list


def is_none(sheet, row, column):
    """ 判断单元格内容是否为空
    :param sheet: <openpyxl.worksheet.worksheet.Worksheet> 需要判断的单元格所在Sheet对象
    :param row: <int> 需要判断单元格所在的行
    :param column: <int> 需要判断单元格所在的列
    :return: <bool> 返回单元格是否为空:True=是,False=否
    """
    return sheet.cell(row=row, column=column).value is None


def row_not_none_col_n(sheet, row, start_column=1, end_column=65536):
    """ 查找某一行中所有不为空的单元格的列坐标
    :param sheet: <openpyxl.worksheet.worksheet.Worksheet> 需要查询的Sheet对象
    :param row: <int> 要查询不为空单元格的行
    :param start_column: <int> 查询不为空的列的范围(从第多少列开始查询)
    :param end_column: <int> 查询不为空的列的范围(查询到多少列为止)
    :return: <list> 不为空的单元格的列坐标,例如: [1,2]
    """
    col_n_list = []
    for j in range(max(1, start_column), min(sheet.max_column, end_column) + 1):
        if not is_none(sheet, row, j):
            col_n_list.append(j)
    return col_n_list


def some_cell_is_none(sheet, row, col_n_list):
    """ 判断一行的中部分列是否存在空值
    :param sheet: <openpyxl.worksheet.worksheet.Worksheet> 需要查询的Sheet对象
    :param row: <int> 需要判断单元格所在的行
    :param col_n_list: <dict> 需要读取的列坐标的列表(此列表可以为None)
    形如: find_some_column的返回结果, 例如: {'平台':1, '目前名称':2}
    :return: <bool> 返回是否存在值为空的单元格:True=是,False=否
    """
    if col_n_list is None:
        return False
    for item in col_n_list:
        if is_none(sheet, row, col_n_list[item]):
            return True
    return False


def cell_value(sheet, row, column, none_value=""):
    """ 读取单元格内容,若单元格为空则返回指定的值
    :param sheet: <openpyxl.worksheet.worksheet.Worksheet> 需要读取的单元格所在Sheet对象
    :param row: <int> 需要读取单元格所在的行
    :param column: <int> 需要读取单元格所在的行
    :param none_value: <object> 若单元格为空时返回的结果
    :return: <object> 单元格内容
    """
    if is_none(sheet, row=row, column=column):
        return none_value
    else:
        return sheet.cell(row=row, column=column).value


def cell_as_str(sheet, row, column):
    """ 读取单元格内容,并将单元格内的值转换为str格式
    :param sheet: <openpyxl.worksheet.worksheet.Worksheet> 需要读取的单元格所在Sheet对象
    :param row: <int> 需要读取单元格所在的行
    :param column: <object> 若单元格为空时返回的结果
    :return: <str> 转换为str格式的单元格内容
    """
    if is_none(sheet, row=row, column=column):
        return ""
    else:
        return str(sheet.cell(row=row, column=column).value)


def cell_as_int(sheet, row, column):
    """ 读取单元格内容,并将单元格内的值转换为int格式(若单元格内容不是数值,则使用正则表达式提取其中的数值部分转换)
    :param sheet: <openpyxl.worksheet.worksheet.Worksheet> 需要读取的单元格所在Sheet对象
    :param row: <int> 需要读取单元格所在的行
    :param column: <object> 若单元格为空时返回的结果
    :return: <int> 转换为int格式的单元格内容
    """
    value = sheet.cell(row=row, column=column).value
    if value is None:
        return 0
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        return int(value)
    if re.search("[0-9]+", str(value)) is not None:
        return int(re.search("[0-9]+", str(value)).group())
    return 0


def cell_as_float(sheet, row, column):
    """ 读取单元格内容,并将单元格内的值转换为float格式(若单元格内容不是数值,则使用正则表达式提取其中的数值部分转换)
    :param sheet: <openpyxl.worksheet.worksheet.Worksheet> 需要读取的单元格所在Sheet对象
    :param row: <int> 需要读取单元格所在的行
    :param column: <object> 若单元格为空时返回的结果
    :return: <float> 转换为float格式的单元格内容
    """
    value = sheet.cell(row=row, column=column).value
    if value is None:
        return float(0)
    if isinstance(value, float):
        return value
    if isinstance(value, int):
        return float(value)
    if re.search("[0-9]+", str(value)) is not None:
        return float(re.search("[0-9]+", str(value)).group())
    if re.search("[0-9]+\.[0-9]+", str(value)) is not None:
        return float(re.search("[0-9]+", str(value)).group())
    return float(0)


def find_cell(sheet, search_value):
    """ 查找指定值的单元格,并返回单元格的行列坐标
    仅会返回查找到的第一个符合查找条件的单元格的行列坐标，在查找时优先按行查找（即完整查完一行后才会开始查找下一行）
    :param sheet: <openpyxl.worksheet.worksheet.Worksheet> 需要查找的单元格的Sheet对象
    :param search_value: <object> 查找单元格的值
    :return: <int> , <int> 返回查找到的单元格的行列坐标,若没有查找到则返回None
    """
    for i in range(1, sheet.max_row + 1):
        for j in range(1, sheet.max_column + 1):
            if sheet.cell(row=i, column=j).value is not None and sheet.cell(row=i, column=j).value == search_value:
                return i, j


def column_total_value(sheet, col_n):
    """ 批量读取Excel表单中的一列数据
    :param sheet: <openpyxl.worksheet.worksheet.Worksheet> 需要读取数据的Sheet对象
    :param col_n: <int> 要读取列的列坐标
    :return: <list> 返回目标列中的数据,按行坐标从小到大排序
    """
    if col_n > sheet.max_column:
        return None
    column_value = []
    for i in range(1, sheet.max_row + 1):
        column_value.append(cell_value(sheet, row=i, column=col_n))
    return column_value


def row_value_by_cn(sheet, row_n, col_n_list):
    """ 批量读取Excel表格中指定行数据中的部分列
    :param sheet: <openpyxl.worksheet.worksheet.Worksheet> 需要读取数据的Sheet对象
    :param row_n: <int> 需要读取的行坐标
    :param col_n_list: <dict> 需要读取的列坐标的列表,形如: find_some_column的返回结果, 例如: {'平台':1, '目前名称':2}
    :return: <list> 返回目标行中的数据,按col_n_list中的顺序
    """
    row_value = []
    for item in col_n_list:
        if col_n_list[item] is not None:
            row_value.append(cell_value(sheet, row=row_n, column=col_n_list[item]))
        else:
            row_value.append(None)
    return row_value


def row_any_value_by_cn(sheet, column_title_row, row_n, column_title_list):
    """ 批量读取Excel表格中指定行数据中的部分列,并选择返回list或该单元格内容
    :param sheet: <openpyxl.worksheet.worksheet.Worksheet> 需要读取数据的Sheet对象
    :param column_title_row: <int> 列名所在的行坐标
    :param row_n: <int> 需要读取的行坐标
    :param column_title_list: <list> 需要读取的列的列名列表
    :return:
    若读取超过一列,则结果按col_n_list中的顺序,返回:[数据,数据,...]
    若读取一列,返回: <object> 该单元格的内容
    """
    sheet_cn_list = find_some_column(sheet, column_title_list, row_n=column_title_row, must_exist=False)
    if len(column_title_list) > 1:
        return row_value_by_cn(sheet, row_n=row_n, col_n_list=sheet_cn_list)
    else:
        if list(sheet_cn_list.keys())[0] is not None:
            return cell_value(sheet, row=row_n, column=sheet_cn_list[list(sheet_cn_list.keys())[0]])
        return None


def get_sheet(sheet, column_title_list, title_rn=1, data_rn=None, classify_column=None, classify_unique=False,
              not_none_column=None):
    """ 批量读取整个Excel表单中的数据
    :param sheet: <openpyxl.worksheet.worksheet.Worksheet> 需要读取数据的Sheet对象
    :param column_title_list: <list> 需要读取的列的列名列表
    :param title_rn: <int> 列名所在的行坐标
    :param data_rn: <int> 开始读取数据的行(默认为列名所在行的后一行)
    :param classify_column: <list/str/none> 如果需要依据某列单元格的值对结果汇总，则填写该列的列名
    :param classify_unique: <bool> 每个分类是否只需要唯一值
    :param not_none_column: <list> 不允许为空值的列名列表,若这些列出现空值则放弃读取该行
    :return: <list/dict> 返回读取Excel表单的结果,按行顺序排序,按col_title_list中的顺序排序;
    具体返回结果会依据classify_column的类型调用get_sheet_by_line,get_sheet_in_classify,get_sheet_some_classify生成
    """
    sheet_nn_cn_list = find_some_column(sheet, basic.not_null_list(not_none_column), row_n=title_rn)
    if data_rn is None:
        data_rn = title_rn + 1
    if isinstance(classify_column, list) and classify_column != []:
        return get_sheet_some_classify(sheet, classify_column, column_title_list, title_rn, data_rn,
                                       classify_unique, sheet_nn_cn_list)
    elif isinstance(classify_column, str):
        return get_sheet_in_classify(sheet, classify_column, column_title_list, title_rn, data_rn,
                                     classify_unique, sheet_nn_cn_list)
    elif classify_column is None or classify_column == []:
        return get_sheet_by_line(sheet, column_title_list, title_rn, data_rn, sheet_nn_cn_list)


def get_sheet_by_line(sheet, column_title_list, title_rn=1, data_rn=2, sheet_nn_cn_list=None):
    """ 批量读取整个Excel表单每行中部分列的数据
    :param sheet: <openpyxl.worksheet.worksheet.Worksheet> 需要读取数据的Sheet对象
    :param column_title_list: <list> 需要读取的列的列名列表
    :param title_rn: <int> 列名所在的行坐标(默认为第1行)
    :param data_rn: <int> 开始读取数据的行(默认为第2行)
    :param sheet_nn_cn_list: <dict> 不允许为空值的列名列表,若这些列出现空值则放弃读取该行
    形如: find_some_column的返回结果, 例如: {'平台':1, '目前名称':2}
    :return: <list> 返回读取Excel表单的结果,按行顺序排序,按col_title_list中的顺序排序
    若读取超过一列,则结果按col_n_list中的顺序,返回:[[数据,数据],[数据,数据]]
    若读取一列,返回:[数据,数据]
    """
    data_list = []
    for i in range(data_rn, sheet.max_row + 1):
        if not some_cell_is_none(sheet, i, sheet_nn_cn_list):
            data_list.append(row_any_value_by_cn(sheet, title_rn, i, column_title_list))
    return data_list


def get_sheet_in_classify(sheet, classify_column, column_title_list, title_rn=1, data_rn=2,
                          classify_unique=False, sheet_nn_cn_list=None):
    """ 批量读取整个Excel表格每行中部分列的数据,并依据其中的某一列对数据进行分类
    :param sheet: <openpyxl.worksheet.worksheet.Worksheet> 需要读取数据的Sheet对象
    :param classify_column: <str> 依据某列单元格的值对结果汇总,则填写该列的列名,不能为空
    :param column_title_list: <list> 需要读取的列的列名列表
    :param title_rn: <int> 列名所在的行坐标(默认为第1行)
    :param data_rn: <int> 开始读取数据的行(默认为第2行)
    :param classify_unique: <bool> 每个分类是否只需要唯一值
    :param sheet_nn_cn_list: <dict> 不允许为空值的列名列表,若这些列出现空值则放弃读取该行
    形如: find_some_column的返回结果, 例如: {'平台':1, '目前名称':2}
    :return: <dict> 返回读取Excel表单的结果,按行顺序排序,按col_title_list中的顺序排序
    若每个分类拥有唯一值,且读取多列,返回: {"分类1":[数据,数据],"分类2":[数据,数据]}
    若每个分类不拥有唯一值,且读取多列,返回: {"分类1":[[数据,数据],[数据,数据]],"分类2":[[数据,数据],[数据,数据]]}
    若每个分类拥有唯一值,且仅读取一列,返回: {"分类1":数据,"分类2":数据}
    若每个分类不拥有唯一值,且仅读取一列,返回: {"分类1":[数据,数据],"分类2":[数据,数据]}
    """
    data_dict = {}
    sheet_cn_classify = find_column(sheet, classify_column, row_n=title_rn)
    for i in range(data_rn, sheet.max_row + 1):
        data_item = row_any_value_by_cn(sheet, title_rn, i, column_title_list)
        data_classify = cell_value(sheet, row=i, column=sheet_cn_classify)
        if classify_unique and data_classify is not None and not some_cell_is_none(sheet, i, sheet_nn_cn_list):
            data_dict[data_classify] = data_item
        if not classify_unique and data_classify is not None and not some_cell_is_none(sheet, i, sheet_nn_cn_list):
            basic.add_dict_to_list(data_dict, data_classify, data_item)
    return data_dict


def get_sheet_some_classify(sheet, classify_column, column_title_list, title_rn=1, data_rn=2,
                            classify_unique=False, sheet_nn_cn_list=None):
    """ 批量读取整个Excel表格每行中部分列的数据,并依据其中的某些列(大于等于一列)对数据进行分类
    :param sheet: <openpyxl.worksheet.worksheet.Worksheet> 需要读取数据的Sheet对象
    :param classify_column: <list> 依据某些列单元格的值对结果汇总,则填写这些列的列名,不能为空
    :param column_title_list: <list> 需要读取的列的列名列表
    :param title_rn: <int> 列名所在的行坐标(默认为第1行)
    :param data_rn: <int> 开始读取数据的行(默认为第2行)
    :param classify_unique: <bool> 每个分类是否只需要唯一值
    :param sheet_nn_cn_list:sheet_nn_cn_list: <dict> 不允许为空值的列名列表,若这些列出现空值则放弃读取该行
    形如: find_some_column的返回结果, 例如: {'平台':1, '目前名称':2}
    :return: <dict> 返回读取Excel表单的结果,按行顺序排序,按col_title_list中的顺序排序
    若每个分类拥有唯一值,且读取多列,以有两层分类为例,返回:
    {"一级分类1":{"二级分类1":[数据,数据],"二级分类2":[数据,数据]},"一级分类2":{"二级分类1":[数据,数据],"二级分类2":[数据,数据]}}
    函数支持有超过两层分类,其中各层分类间的数据结构与以上结构相似,内层数据结构与get_sheet_in_classify的返回结果类似
    """
    data_dict = {}
    sheet_clfy_cn_list = find_some_column(sheet, classify_column, row_n=title_rn)
    for i in range(data_rn, sheet.max_row + 1):
        if not some_cell_is_none(sheet, i, sheet_clfy_cn_list) and not some_cell_is_none(sheet, i, sheet_nn_cn_list):
            data_item = row_any_value_by_cn(sheet, title_rn, i, column_title_list)
            data_classify = data_dict
            for j in range(len(classify_column) - 1):
                classify_name = cell_value(sheet, row=i, column=sheet_clfy_cn_list[classify_column[j]])
                if classify_name not in data_classify:
                    data_classify[classify_name] = {}
                data_classify = data_classify[classify_name]
            last_classify_name = cell_value(sheet, row=i,
                                            column=sheet_clfy_cn_list[classify_column[len(classify_column) - 1]])
            if last_classify_name != "":
                if classify_unique:
                    data_classify[last_classify_name] = data_item
                else:
                    basic.add_dict_to_list(data_classify, last_classify_name, data_item)
    return data_dict


def write_row(sheet, row, value_list):
    """ 批量写入一行数据
    :param sheet: <openpyxl.worksheet.worksheet.Worksheet> 需要写入数据的Sheet对象
    :param row: <int> 需要写入的行
    :param value_list: <list/None> 需要写入到目标行的值
    :return:隐含返回值,结果更新于[sheet]
    """
    if value_list is not None:
        for i in range(len(value_list)):
            sheet.cell(row=row, column=i + 1).value = value_list[i]


def write_column(sheet, column, value_list):
    """ 批量写入一列数据
    :param sheet: <openpyxl.worksheet.worksheet.Worksheet> 需要写入数据的Sheet对象
    :param column: <int> 需要写入的列
    :param value_list: <list/None> 需要写入到目标列的值
    :return:隐含返回值,结果更新于[sheet]
    """
    if value_list is not None:
        for i in range(len(value_list)):
            sheet.cell(row=i + 1, column=column).value = value_list[i]


def write_sheet_cls_list(sheet, title, data, classify_index):
    """ 批量写入数据到表单(key为分类,value为该分类多条记录list的list,的dict数据)
    例如:{"分类1":[[列1,列2],[列1,列2]],"分类2":[[列1,列2],[列1,列2]]}
    :param sheet: <openpyxl.worksheet.worksheet.Worksheet> 需要设置行高的Sheet对象
    :param title: <list> 写入数据的标题
    :param data: <dict> 写入数据的内容
    :param classify_index:分类字段添加到每行数据中的列坐标(从0开始)
    :return:隐含返回值,结果更新于[sheet]
    """
    write_row(sheet, 1, title)
    sheet_rn = 2
    for classify in data:
        for item in copy.deepcopy(data[classify]):
            item.insert(classify_index, classify)
            write_row(sheet, sheet_rn, item)
            sheet_rn += 1


def write_sheet_cls_item(sheet, title, data, classify_index=0):
    """ 批量写入数据到表单(key为分类,value为单条记录list,的dict数据)
    例如:{"分类1":[列1,列2],"分类2":[列1,列2]}
    :param sheet: <openpyxl.worksheet.worksheet.Worksheet> 写出数据的Sheet对象
    :param title: <list> 写入数据的标题
    :param data: <dict> 写入数据的内容
    :param classify_index:分类字段添加到每行数据中的列坐标(从0开始)
    :return:隐含返回值,结果更新于[sheet]
    """
    write_row(sheet, 1, title)
    sheet_rn = 2
    for classify in data:
        item = copy.deepcopy(data[classify])
        item.insert(classify_index, classify)
        write_row(sheet, sheet_rn, item)
        sheet_rn += 1


def get_border(sheet, row, column, dTop=True, dRight=True, dBottom=True, dLeft=True):
    """ 获取单元格边框样式
    :param sheet: <openpyxl.worksheet.worksheet.Worksheet> 需要获取样式的Sheet对象
    :param row: <int> 目标单元格的行坐标
    :param column: <int> 目标单元格的列坐标
    :param dTop: <bool> 是否提取该单元格上方的边框样式
    :param dRight: <bool> 是否提取该单元格右方的边框样式
    :param dBottom: <bool> 是否提取该单元格下方的边框样式
    :param dLeft: <bool> 是否提取该单元格左方的边框样式
    :return: <openpyxl.styles.Border> 目标单元格的边框样式
    """
    tBorder = sheet.cell(row=row, column=column).border
    side = {"top": Side(), "right": Side(), "bottom": Side(), "left": Side()}
    if dTop:
        side["top"] = tBorder.top
    if dRight:
        side["right"] = tBorder.right
    if dBottom:
        side["bottom"] = tBorder.bottom
    if dLeft:
        side["left"] = tBorder.left
    return Border(top=side["top"], right=side["right"], bottom=side["bottom"], left=side["left"])


def get_height(sheet, row_num_list):
    """ 获取行高列表
    :param sheet: <openpyxl.worksheet.worksheet.Worksheet> 需要获取样式的Sheet对象
    :param row_num_list: <list> 行坐标列表
    :return: <list> 行高列表
    """
    height = []
    for rn in row_num_list:
        height.append(sheet.row_dimensions[rn].height)
    return height


def set_column_width(sheet, width_list, start_column=1):
    """ 批量设置列宽
    :param sheet: <openpyxl.worksheet.worksheet.Worksheet> 需要设置列宽的Sheet对象
    :param width_list: <list> 要设置的各列的列宽列表
    :param start_column: <int> 开始设置列宽的列(默认从第1列的列宽开始设置)
    :return: <None>
    """
    for i in range(len(width_list)):
        sheet.column_dimensions[get_column_letter(start_column + i)].width = width_list[i]


def set_row_height(sheet, height_list, start_row=1):
    """ 批量设置行高
    :param sheet: <openpyxl.worksheet.worksheet.Worksheet> 需要设置行高的Sheet对象
    :param height_list: <list> 要设置的各行的行高列表
    :param start_row: <int> 开始设置行高的行(默认从第1行的行高开始设置)
    :return: <None>
    """
    for i in range(len(height_list)):
        sheet.row_dimensions[start_row + i].height = height_list[i]
